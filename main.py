from pathlib import Path

import numpy as np
import openpyxl as excel
from openpyxl.styles.alignment import Alignment
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
from openpyxl.cell.cell import Cell

import tools


@tools.timer
def main():
    path: Path = tools.get_file_path()

    workbook: Workbook = excel.load_workbook(path)
    alignment = Alignment(horizontal='center',vertical='center')
    for _name in workbook.sheetnames:
        sheet: Worksheet = workbook[_name]
        workbook.create_sheet(_name + "-avg")
        sheet_avg: Worksheet = workbook[_name + "-avg"]
        for i in range(len(sheet.row_dimensions)):
            row_index = tools.to_row(i)
            sheet_avg.row_dimensions[row_index].height = sheet.row_dimensions[row_index].height
        for i in range(len(sheet.column_dimensions)):
            column_index = tools.to_column(i)
            sheet_avg.column_dimensions[column_index].width = sheet.column_dimensions[column_index].width

        tools.copy_cell(sheet["A:C"], sheet_avg)

        group_num: int = int((sheet.max_column - 3) / 4)
        for i in range(group_num):
            src_start_point = ord("D") + i * 4
            tar_start_point = ord("D") + i
            # 复制表头
            tools.copy_cell(sheet[f"{chr(src_start_point)}1"], sheet_avg[f"{chr(tar_start_point)}1"])

            # 获取数值
            coordinate_src = f"{chr(src_start_point)}2:{chr(src_start_point + 3)}{sheet.max_row + 1}"
            nums_list = tools.get_value(sheet[coordinate_src], 0)

            # 取均值并赋值
            for j in range(len(nums_list)):
                nums = []
                for num in nums_list[j]:
                    if isinstance(num, int | float):
                        nums.append(num)

                tar_cell: Cell = sheet_avg[f"{chr(tar_start_point)}{2 + j}"]
                tar_cell.number_format = "0.00"
                tar_cell.alignment = alignment
                match len(nums):
                    case 1:
                        tar_cell.value = nums[0]
                    case 2 | 3:
                        avg = np.mean(nums)
                        tar_cell.value = avg
                    case 4:
                        avg = np.mean(nums[0:2])
                        tar_cell.value = avg

    new_path = path.parent / f"{path.stem}_1{path.suffix}"
    workbook.save(new_path)


if __name__ == '__main__':
    main()
